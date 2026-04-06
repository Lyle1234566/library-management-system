from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - guarded at runtime
    load_workbook = None

from .models import TeacherRecord

TEACHER_TEMPLATE_COLUMNS = (
    'staff_id',
    'full_name',
    'school_email',
    'department',
    'is_active_for_registration',
    'notes',
)

HEADER_ALIASES = {
    'staff_id': {
        'staff_id',
        'staffid',
        'faculty_id',
        'facultyid',
        'employee_id',
        'employeeid',
        'teacher_id',
        'teacherid',
        'id_number',
    },
    'full_name': {'full_name', 'fullname', 'teacher_name', 'teachername', 'name'},
    'school_email': {'school_email', 'schoolemail', 'teacher_email', 'teacheremail', 'email', 'email_address'},
    'department': {'department', 'college', 'division', 'unit'},
    'is_active_for_registration': {
        'is_active_for_registration',
        'active_for_registration',
        'activeforregistration',
        'is_active',
        'active',
        'enabled',
    },
    'notes': {'notes', 'remarks', 'comment', 'comments'},
}

HEADER_LOOKUP = {
    alias: canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}
SUPPORTED_IMPORT_EXTENSIONS = {'.csv', '.xlsx'}


class TeacherImportError(ValueError):
    pass


@dataclass(frozen=True)
class TeacherImportResult:
    created_count: int
    updated_count: int
    skipped_count: int
    skipped_rows: list[str]


def parse_bool(value: str | None, default: bool = True) -> bool:
    if value is None or value == '':
        return default
    return str(value).strip().lower() in {'1', 'true', 'yes', 'y', 'on'}


def normalize_header(value: object) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')


def resolve_header(value: object) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None
    return HEADER_LOOKUP.get(normalized, normalized if normalized in TEACHER_TEMPLATE_COLUMNS else None)


def stringify_cell(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _iter_normalized_rows(headers: object, row_iter, start_row_number: int = 2):
    canonical_headers = [resolve_header(header) for header in headers]
    if 'staff_id' not in canonical_headers:
        raise TeacherImportError('File must include a staff_id or faculty_id column.')

    for row_number, raw_row in enumerate(row_iter, start=start_row_number):
        values = list(raw_row or [])
        if not any(stringify_cell(value) for value in values):
            continue

        normalized_row: dict[str, str] = {}
        for index, canonical_header in enumerate(canonical_headers):
            if canonical_header is None:
                continue
            cell_value = values[index] if index < len(values) else ''
            normalized_row[canonical_header] = stringify_cell(cell_value)

        yield row_number, normalized_row


def _import_rows(rows) -> TeacherImportResult:
    created_count = 0
    updated_count = 0
    skipped_rows: list[str] = []

    for row_number, row in rows:
        staff_id = stringify_cell(row.get('staff_id')).upper()
        if not staff_id:
            skipped_rows.append(f'Row {row_number}: missing staff_id.')
            continue

        defaults = {
            'full_name': stringify_cell(row.get('full_name')),
            'school_email': stringify_cell(row.get('school_email')).lower(),
            'department': stringify_cell(row.get('department')),
            'is_active_for_registration': parse_bool(
                row.get('is_active_for_registration'),
                default=True,
            ),
            'notes': stringify_cell(row.get('notes')),
        }

        _, created = TeacherRecord.objects.update_or_create(
            staff_id=staff_id,
            defaults=defaults,
        )
        if created:
            created_count += 1
        else:
            updated_count += 1

    return TeacherImportResult(
        created_count=created_count,
        updated_count=updated_count,
        skipped_count=len(skipped_rows),
        skipped_rows=skipped_rows[:20],
    )


def import_teacher_csv_text(csv_text: str) -> TeacherImportResult:
    reader = csv.reader(StringIO(csv_text))
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise TeacherImportError('CSV file is empty.') from exc

    return _import_rows(_iter_normalized_rows(headers, reader))


def import_teacher_csv_file(uploaded_file) -> TeacherImportResult:
    raw_content = uploaded_file.read()
    if isinstance(raw_content, bytes):
        try:
            csv_text = raw_content.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise TeacherImportError('CSV must be UTF-8 encoded.') from exc
    else:
        csv_text = str(raw_content)

    return import_teacher_csv_text(csv_text)


def import_teacher_xlsx_file(uploaded_file) -> TeacherImportResult:
    if load_workbook is None:
        raise TeacherImportError('Excel import is unavailable because openpyxl is not installed.')

    raw_content = uploaded_file.read()
    try:
        workbook = load_workbook(filename=BytesIO(raw_content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - library-specific errors vary
        raise TeacherImportError('Excel file could not be read.') from exc

    try:
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        headers = next(row_iter, None)
        if headers is None:
            raise TeacherImportError('Excel file is empty.')
        return _import_rows(_iter_normalized_rows(headers, row_iter))
    finally:
        workbook.close()


def import_teacher_file(uploaded_file) -> TeacherImportResult:
    suffix = Path(getattr(uploaded_file, 'name', '') or '').suffix.lower()
    if suffix == '.xlsx':
        return import_teacher_xlsx_file(uploaded_file)
    if suffix in {'', '.csv'}:
        return import_teacher_csv_file(uploaded_file)
    raise TeacherImportError('Unsupported file type. Upload a CSV or Excel (.xlsx) file.')


def get_teacher_summary() -> dict[str, int | str | list[str] | None]:
    total_records = TeacherRecord.objects.count()
    active_records = TeacherRecord.objects.filter(is_active_for_registration=True).count()
    last_updated = TeacherRecord.objects.order_by('-updated_at').values_list('updated_at', flat=True).first()

    return {
        'total_records': total_records,
        'active_records': active_records,
        'inactive_records': total_records - active_records,
        'last_updated_at': last_updated.isoformat() if last_updated else None,
        'template_columns': list(TEACHER_TEMPLATE_COLUMNS),
        'supported_extensions': sorted(SUPPORTED_IMPORT_EXTENSIONS),
    }
