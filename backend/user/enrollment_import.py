from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - guarded at runtime
    load_workbook = None

from .models import EnrollmentRecord

ENROLLMENT_TEMPLATE_COLUMNS = (
    'student_id',
    'full_name',
    'school_email',
    'program',
    'year_level',
    'academic_term',
    'is_currently_enrolled',
    'notes',
)

HEADER_ALIASES = {
    'student_id': {'student_id', 'studentid', 'student_no', 'student_number', 'studentnumber', 'id_number'},
    'full_name': {'full_name', 'fullname', 'student_name', 'studentname', 'name'},
    'school_email': {'school_email', 'schoolemail', 'student_email', 'studentemail', 'email', 'email_address'},
    'program': {'program', 'course'},
    'year_level': {'year_level', 'yearlevel', 'year'},
    'academic_term': {'academic_term', 'academicterm', 'term', 'semester', 'school_year', 'schoolyear'},
    'is_currently_enrolled': {
        'is_currently_enrolled',
        'currently_enrolled',
        'currentlyenrolled',
        'is_enrolled',
        'isenrolled',
        'enrolled',
        'active',
    },
    'notes': {'notes', 'remarks', 'comment', 'comments'},
}

HEADER_LOOKUP = {
    alias: canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}
SUPPORTED_IMPORT_EXTENSIONS = {'.csv', '.xlsx'}


class EnrollmentImportError(ValueError):
    pass


@dataclass(frozen=True)
class EnrollmentImportResult:
    created_count: int
    updated_count: int
    skipped_count: int
    skipped_rows: list[str]


def parse_bool(value: str | None, default: bool = True) -> bool:
    if value is None or value == '':
        return default
    return str(value).strip().lower() in {'1', 'true', 'yes', 'y', 'on'}


def normalize_header(value: object) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')


def resolve_header(value: object) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None
    return HEADER_LOOKUP.get(normalized, normalized if normalized in ENROLLMENT_TEMPLATE_COLUMNS else None)


def stringify_cell(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _iter_normalized_rows(headers: object, row_iter, start_row_number: int = 2):
    canonical_headers = [resolve_header(header) for header in headers]
    if 'student_id' not in canonical_headers:
        raise EnrollmentImportError('File must include a student_id column.')

    for row_number, raw_row in enumerate(row_iter, start=start_row_number):
        values = list(raw_row or [])
        if not any(stringify_cell(value) for value in values):
            continue

        normalized_row: dict[str, str] = {}
        for index, canonical_header in enumerate(canonical_headers):
            if canonical_header is None:
                continue
            cell_value = values[index] if index < len(values) else ''
            normalized_row[canonical_header] = stringify_cell(cell_value)

        yield row_number, normalized_row


def _import_rows(rows, fallback_term: str = '') -> EnrollmentImportResult:
    created_count = 0
    updated_count = 0
    skipped_rows: list[str] = []

    for row_number, row in rows:
        student_id = stringify_cell(row.get('student_id')).upper()
        if not student_id:
            skipped_rows.append(f'Row {row_number}: missing student_id.')
            continue

        defaults = {
            'full_name': stringify_cell(row.get('full_name')),
            'school_email': stringify_cell(row.get('school_email')).lower(),
            'program': stringify_cell(row.get('program')),
            'year_level': stringify_cell(row.get('year_level')),
            'academic_term': stringify_cell(row.get('academic_term') or fallback_term),
            'is_currently_enrolled': parse_bool(row.get('is_currently_enrolled'), default=True),
            'notes': stringify_cell(row.get('notes')),
        }

        _, created = EnrollmentRecord.objects.update_or_create(
            student_id=student_id,
            defaults=defaults,
        )
        if created:
            created_count += 1
        else:
            updated_count += 1

    return EnrollmentImportResult(
        created_count=created_count,
        updated_count=updated_count,
        skipped_count=len(skipped_rows),
        skipped_rows=skipped_rows[:20],
    )


def import_enrollment_csv_text(csv_text: str, fallback_term: str = '') -> EnrollmentImportResult:
    reader = csv.reader(StringIO(csv_text))
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise EnrollmentImportError('CSV file is empty.') from exc

    return _import_rows(
        _iter_normalized_rows(headers, reader),
        fallback_term=fallback_term,
    )


def import_enrollment_csv_file(uploaded_file, fallback_term: str = '') -> EnrollmentImportResult:
    raw_content = uploaded_file.read()
    if isinstance(raw_content, bytes):
        try:
            csv_text = raw_content.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise EnrollmentImportError('CSV must be UTF-8 encoded.') from exc
    else:
        csv_text = str(raw_content)

    return import_enrollment_csv_text(csv_text, fallback_term=fallback_term)


def import_enrollment_xlsx_file(uploaded_file, fallback_term: str = '') -> EnrollmentImportResult:
    if load_workbook is None:
        raise EnrollmentImportError('Excel import is unavailable because openpyxl is not installed.')

    raw_content = uploaded_file.read()
    try:
        workbook = load_workbook(filename=BytesIO(raw_content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - library-specific errors vary
        raise EnrollmentImportError('Excel file could not be read.') from exc

    try:
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        headers = next(row_iter, None)
        if headers is None:
            raise EnrollmentImportError('Excel file is empty.')
        return _import_rows(
            _iter_normalized_rows(headers, row_iter),
            fallback_term=fallback_term,
        )
    finally:
        workbook.close()


def import_enrollment_file(uploaded_file, fallback_term: str = '') -> EnrollmentImportResult:
    suffix = Path(getattr(uploaded_file, 'name', '') or '').suffix.lower()
    if suffix == '.xlsx':
        return import_enrollment_xlsx_file(uploaded_file, fallback_term=fallback_term)
    if suffix in {'', '.csv'}:
        return import_enrollment_csv_file(uploaded_file, fallback_term=fallback_term)
    raise EnrollmentImportError('Unsupported file type. Upload a CSV or Excel (.xlsx) file.')


def import_enrollment_csv_path(csv_path: Path, fallback_term: str = '') -> EnrollmentImportResult:
    try:
        csv_text = csv_path.read_text(encoding='utf-8-sig')
    except FileNotFoundError as exc:
        raise EnrollmentImportError(f'CSV file not found: {csv_path}') from exc

    return import_enrollment_csv_text(csv_text, fallback_term=fallback_term)


def get_enrollment_summary() -> dict[str, int | str | list[str] | None]:
    total_records = EnrollmentRecord.objects.count()
    active_records = EnrollmentRecord.objects.filter(is_currently_enrolled=True).count()
    latest_term = (
        EnrollmentRecord.objects.exclude(academic_term='')
        .order_by('-updated_at')
        .values_list('academic_term', flat=True)
        .first()
    )
    last_updated = EnrollmentRecord.objects.order_by('-updated_at').values_list('updated_at', flat=True).first()

    return {
        'total_records': total_records,
        'active_records': active_records,
        'inactive_records': total_records - active_records,
        'latest_term': latest_term or None,
        'last_updated_at': last_updated.isoformat() if last_updated else None,
        'template_columns': list(ENROLLMENT_TEMPLATE_COLUMNS),
        'supported_extensions': sorted(SUPPORTED_IMPORT_EXTENSIONS),
    }
